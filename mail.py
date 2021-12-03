from openpyxl import load_workbook


class Mail:
    """
     把客户信息拼接到话术里
    """
    def __init__(self):
        self.wb = load_workbook('拒收需发邮件.xlsx')
        self.ws = self.wb.active
        self.number = 2

    def read_track_table(self):
        """
        :return:
        """
        rows = []
        for row in self.ws.iter_rows():
            rows.append(row)
        for x in range(1, len(rows)):
            mails = []
            number = str(rows[x][0].value)
            order = str(rows[x][1].value)
            name = str(rows[x][2].value)
            contact_email = str(rows[x][3].value)
            nat = str(rows[x][4].value)
            reason = str(rows[x][5].value)
            mails.append(number)
            mails.append(order)
            mails.append(name)
            mails.append(contact_email)
            mails.append(nat)
            mails.append(reason)
            yield mails

    def read_txt(self):
        for na in self.read_track_table():
            with open(f'national_speech/{na[4]}.txt', 'r', encoding='utf-8') as f:
                words = f.read().format(na[2], na[0], na[1])
                self.ws[f'G{self.number}'] = words
                self.wb.save('拒收需发邮件.xlsx')
                self.number += 1
